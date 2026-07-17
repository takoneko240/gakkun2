import os

import openpyxl

SAMPLE_ROWS = [
    ("希望", "きぼう"),
    ("愛する", "あいする"),
    ("学校", "がっこう"),
    ("勉強", "べんきょう"),
    ("友達", "ともだち"),
]


def ensure_sample_file(path):
    if os.path.exists(path):
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["漢字", "読み"])
    for kanji, yomi in SAMPLE_ROWS:
        ws.append([kanji, yomi])
    wb.save(path)


def load_kanji_list(path):
    if not os.path.exists(path):
        return []

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    entries = []
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        kanji, yomi = row[0], row[1]
        if kanji is None or yomi is None:
            continue
        kanji = str(kanji).strip()
        yomi = str(yomi).strip()
        if kanji and yomi:
            entries.append((kanji, yomi))

    wb.close()
    return entries
